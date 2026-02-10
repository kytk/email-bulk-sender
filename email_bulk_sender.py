import smtplib
import csv
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.header import Header
from email.utils import formataddr
from email import encoders
import time
from getpass import getpass
import os
from urllib.parse import quote
import mimetypes
import chardet
import argparse
import json
from pathlib import Path
from typing import Dict, Any, Optional

# ==================== 内部モジュール定義 ====================
# i18n.py と config.py が存在しない場合に使用される内部定義

# --- I18n (国際化) モジュール ---
class I18n:
    """国際化クラス（内部定義版）"""

    TEXTS = {
        'ja': {
            'cli_title': '=== メール一括送信ツール ===',
            'cli_smtp_server': 'SMTPサーバー (例: smtp.gmail.com)',
            'cli_smtp_port': 'SMTPポート (デフォルト: 587)',
            'cli_email_address': '送信元メールアドレス',
            'cli_email_password': 'メールパスワード',
            'cli_sender_name': '送信元表示名 (不要ならEnter)',
            'cli_csv_file': '受信者リストCSVファイル (デフォルト: list.csv)',
            'cli_template_file': 'メールテンプレートファイル (デフォルト: body.txt)',
            'cli_cc': 'CC (複数の場合はカンマ区切り、不要ならEnter)',
            'cli_bcc': 'BCC (複数の場合はカンマ区切り、不要ならEnter)',
            'cli_reply_to': 'Reply-To (不要ならEnter)',
            'cli_confirm_header': '\n=== 送信内容確認 ===',
            'cli_confirm_send': '\n送信を開始しますか？ (yes/no)',
            'cli_cancelled': '送信をキャンセルしました',
            'preview_subject': '件名: {0}',
            'preview_recipients': '送信先: {0}件',
            'preview_sender': '送信元: {0}',
            'preview_cc': 'CC: {0}',
            'preview_bcc': 'BCC: {0}',
            'preview_reply_to': 'Reply-To: {0}',
            'preview_attachments': '添付ファイル: {0}',
            'send_success': '[{0}/{1}] 送信成功: {2} {3} ({4})',
            'send_failed': '[{0}/{1}] 送信失敗: {2} {3} ({4}) - {5}',
            'send_complete': '送信完了: 成功 {0}件, 失敗 {1}件',
        },
        'en': {
            'cli_title': '=== Email Bulk Sender ===',
            'cli_smtp_server': 'SMTP Server (e.g., smtp.gmail.com)',
            'cli_smtp_port': 'SMTP Port (default: 587)',
            'cli_email_address': 'Sender Email Address',
            'cli_email_password': 'Email Password',
            'cli_sender_name': 'Sender Display Name (press Enter to skip)',
            'cli_csv_file': 'Recipient List CSV File (default: list.csv)',
            'cli_template_file': 'Email Template File (default: body.txt)',
            'cli_cc': 'CC (comma-separated for multiple, press Enter to skip)',
            'cli_bcc': 'BCC (comma-separated for multiple, press Enter to skip)',
            'cli_reply_to': 'Reply-To (press Enter to skip)',
            'cli_confirm_header': '\n=== Confirm Sending Details ===',
            'cli_confirm_send': '\nStart sending? (yes/no)',
            'cli_cancelled': 'Sending cancelled',
            'preview_subject': 'Subject: {0}',
            'preview_recipients': 'Recipients: {0}',
            'preview_sender': 'Sender: {0}',
            'preview_cc': 'CC: {0}',
            'preview_bcc': 'BCC: {0}',
            'preview_reply_to': 'Reply-To: {0}',
            'preview_attachments': 'Attachments: {0}',
            'send_success': '[{0}/{1}] Success: {2} {3} ({4})',
            'send_failed': '[{0}/{1}] Failed: {2} {3} ({4}) - {5}',
            'send_complete': 'Sending complete: {0} succeeded, {1} failed',
        }
    }

    def __init__(self, lang=None):
        import locale
        if lang is None:
            try:
                system_locale = locale.getlocale()[0]
                lang = 'ja' if system_locale and system_locale.startswith('ja') else 'en'
            except:
                lang = 'en'
        self.lang = lang if lang in ['ja', 'en'] else 'en'

    def get(self, key, *args):
        text = self.TEXTS.get(self.lang, {}).get(key, key)
        if args:
            return text.format(*args)
        return text

    def get_language(self):
        return self.lang

class InternalConfigManager:
    """設定ファイル管理クラス（内部定義版）"""

    CONFIG_VERSION = "2.0"

    def __init__(self, config_type: str = "email"):
        self.config_type = config_type
        if config_type == "gmail":
            self.config_dir = Path.home() / ".gmail_bulk_sender"
        else:
            self.config_dir = Path.home() / ".email_bulk_sender"
        self.config_file = self.config_dir / "config.json"

    def get_default_config(self) -> Dict[str, Any]:
        base_config = {
            "version": self.CONFIG_VERSION,
            "sender": {"email_address": "", "display_name": ""},
            "files": {"csv_file": "", "template_file": "", "attachments": []},
            "email_options": {"cc": "", "bcc": "", "reply_to": "", "send_delay": 5},
            "ui": {"language": "ja"}
        }
        if self.config_type == "email":
            base_config["smtp"] = {"server": "", "port": 587}
        return base_config

    def load_config(self) -> Optional[Dict[str, Any]]:
        if not self.config_file.exists():
            return None
        try:
            with open(self.config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
            if "version" not in config:
                config["version"] = self.CONFIG_VERSION
            return config
        except (json.JSONDecodeError, IOError) as e:
            print(f"Error loading config file: {e}")
            return None

    def save_config(self, config: Dict[str, Any]) -> bool:
        try:
            self.config_dir.mkdir(parents=True, exist_ok=True)
            if "version" not in config:
                config["version"] = self.CONFIG_VERSION
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            os.chmod(self.config_file, 0o600)
            return True
        except (IOError, OSError) as e:
            print(f"Error saving config file: {e}")
            return False

    def get_config_path(self) -> str:
        return str(self.config_file)

# --- 条件付きインポート ---
# 外部ファイルが存在する場合はそれを使用、なければ内部定義を使用
try:
    from i18n import get_i18n
except ImportError:
    def get_i18n(lang=None):
        return I18n(lang)

try:
    from config import ConfigManager
except ImportError:
    ConfigManager = InternalConfigManager

# =======================================================

# Excel対応
try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# ==================== 設定セクション ====================
# ここで送信元情報とSMTPサーバー設定をしてください

# SMTPサーバー（空文字列の場合は実行時に入力を求めます）
DEFAULT_SMTP_SERVER = ""  # 例: "smtp.example.com" or "smtp.gmail.com"

# SMTPポート番号（空文字列の場合は実行時に入力を求めます。デフォルト: 587）
DEFAULT_SMTP_PORT = ""  # 例: "587" (TLS), "465" (SSL), "25" (非暗号化)

# メールアドレス（空文字列の場合は実行時に入力を求めます）
DEFAULT_EMAIL_ADDRESS = ""  # 例: "your.email@example.com"

# メールパスワード（空文字列の場合は実行時に入力を求めます）
# セキュリティ上、ここには記載せず実行時に入力することを推奨
DEFAULT_EMAIL_PASSWORD = ""

# 送信元表示名（空文字列の場合はメールアドレスのみ表示）
SENDER_DISPLAY_NAME = ""  # 例: "株式会社サンプル 営業部"

# 受信者リストCSVファイル（空文字列の場合はデフォルト: list.csv）
DEFAULT_CSV_FILE = ""  # 例: "recipients.csv"

# メールテンプレートファイル（空文字列の場合はデフォルト: body.txt）
DEFAULT_TEMPLATE_FILE = ""  # 例: "email_template.txt"

# CC（Noneの場合は実行時に入力を求め、""の場合は不要としてスキップ）
DEFAULT_CC = None  # 例: "cc@example.com" または "" (不要な場合)

# BCC（Noneの場合は実行時に入力を求め、""の場合は不要としてスキップ）
DEFAULT_BCC = None  # 例: "bcc@example.com" または "" (不要な場合)

# Reply-To（Noneの場合は実行時に入力を求め、""の場合は不要としてスキップ）
DEFAULT_REPLY_TO = None  # 例: "reply@example.com" または "" (不要な場合)

# 添付ファイル（空文字列の場合は実行時に入力を求めます）
# 複数のファイルを添付する場合はカンマ区切りで指定
DEFAULT_ATTACHMENTS = ""  # 例: "file1.pdf,file2.docx"

# メール送信間隔（秒）- スパム扱いを避けるための遅延時間
# 推奨値: 少量(~50通)=3-5秒, 中量(50-100通)=5-10秒, 大量(100通以上)=10秒以上
DEFAULT_SEND_DELAY = 5  # デフォルト: 5秒

# =======================================================

class EmailBulkSender:
    def __init__(self, email_address, email_password, smtp_server, smtp_port, sender_display_name=""):
        """
        メール一斉送信クラスの初期化

        Args:
            email_address: 送信元メールアドレス
            email_password: メールパスワード
            smtp_server: SMTPサーバーアドレス
            smtp_port: SMTPポート番号
            sender_display_name: 送信元表示名（オプション）
        """
        self.email_address = email_address
        self.email_password = email_password
        self.smtp_server = smtp_server
        self.smtp_port = smtp_port
        self.sender_display_name = sender_display_name
    
    def _get_safe_local_hostname(self) -> str:
        """SMTP EHLO用の安全なローカルホスト名を取得（非ASCIIコンピュータ名対応）"""
        import socket
        fqdn = socket.getfqdn()
        try:
            fqdn.encode('ascii')
            return fqdn
        except UnicodeEncodeError:
            return 'localhost'

    def read_recipients(self, csv_file, i18n=None):
        """
        CSVまたはExcelファイルから受信者リストを読み込む（文字コード自動検出）

        Args:
            csv_file: CSVまたはExcelファイルのパス（所属,氏名,メールアドレスの形式）
            i18n: 国際化インスタンス

        Returns:
            受信者の辞書リスト [{'affiliation': '株式会社ABC', 'name': '山田太郎', 'email': 'yamada@example.com'}, ...]
        """
        # ファイル拡張子で判定
        file_ext = os.path.splitext(csv_file)[1].lower()

        if file_ext == '.xlsx':
            # Excelファイルの場合
            if not EXCEL_SUPPORT:
                error_msg = "Excel file support requires openpyxl. Install it with: pip install openpyxl" if i18n is None else i18n.t("error_excel_support")
                raise ImportError(error_msg)

            if i18n:
                print(f"Reading Excel file: {csv_file}")
            else:
                print(f"Excelファイルを読み込み中: {csv_file}")

            return self._read_recipients_from_excel(csv_file, i18n)
        else:
            # CSVファイルの場合
            return self._read_recipients_from_csv(csv_file, i18n)

    def _read_recipients_from_csv(self, csv_file, i18n=None):
        """
        CSVファイルから受信者リストを読み込む（文字コード自動検出）

        Args:
            csv_file: CSVファイルのパス
            i18n: 国際化インスタンス

        Returns:
            受信者の辞書リスト
        """
        # ファイルの文字コードを自動検出
        with open(csv_file, 'rb') as f:
            raw_data = f.read()
            detected = chardet.detect(raw_data)
            encoding = detected['encoding']
            confidence = detected['confidence']
            # i18nがない場合はデフォルトメッセージ（後方互換性のため）
            if i18n:
                print(f"CSV encoding: {encoding} (confidence: {confidence:.2%})")
            else:
                print(f"CSVファイルの文字コード: {encoding} (信頼度: {confidence:.2%})")

        # 検出された文字コードでファイルを読み込む（newline=''で改行コードを適切に処理）
        recipients = []
        with open(csv_file, 'r', encoding=encoding, newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                # CSVのカラム名に応じて調整
                affiliation_key = '所属' if '所属' in row else 'affiliation'
                name_key = '氏名' if '氏名' in row else 'name'
                email_key = 'メールアドレス' if 'メールアドレス' in row else 'email'

                recipients.append({
                    'affiliation': row[affiliation_key].strip(),
                    'name': row[name_key].strip(),
                    'email': row[email_key].strip()
                })
        return recipients

    def _read_recipients_from_excel(self, xlsx_file, i18n=None):
        """
        Excelファイルから受信者リストを読み込む

        Args:
            xlsx_file: Excelファイルのパス
            i18n: 国際化インスタンス

        Returns:
            受信者の辞書リスト
        """
        workbook = load_workbook(xlsx_file, read_only=True)
        sheet = workbook.active

        recipients = []
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) == 0:
            return recipients

        # ヘッダー行を取得
        header = rows[0]

        # カラムインデックスを特定
        affiliation_key = '所属' if '所属' in header else 'affiliation'
        name_key = '氏名' if '氏名' in header else 'name'
        email_key = 'メールアドレス' if 'メールアドレス' in header else 'email'

        try:
            affiliation_idx = header.index(affiliation_key)
            name_idx = header.index(name_key)
            email_idx = header.index(email_key)
        except ValueError as e:
            error_msg = f"Required columns not found in Excel file. Expected: {affiliation_key}, {name_key}, {email_key}" if i18n is None else i18n.t("error_excel_columns")
            raise ValueError(error_msg) from e

        # データ行を読み込む
        for row in rows[1:]:
            if row[email_idx]:  # メールアドレスがある行のみ
                recipients.append({
                    'affiliation': str(row[affiliation_idx]).strip() if row[affiliation_idx] else '',
                    'name': str(row[name_idx]).strip() if row[name_idx] else '',
                    'email': str(row[email_idx]).strip()
                })

        workbook.close()
        return recipients
    
    def read_email_template(self, template_file, i18n=None):
        """
        メールテンプレートを読み込む（件名と本文を分離、文字コード自動検出）

        Args:
            template_file: テンプレートファイルのパス
            i18n: 国際化インスタンス

        Returns:
            (subject, body) のタプル
        """
        # ファイルの文字コードを自動検出
        with open(template_file, 'rb') as f:
            raw_data = f.read()
            detected = chardet.detect(raw_data)
            encoding = detected['encoding']
            confidence = detected['confidence']
            # i18nがない場合はデフォルトメッセージ（後方互換性のため）
            if i18n:
                print(f"Template encoding: {encoding} (confidence: {confidence:.2%})")
            else:
                print(f"テンプレートファイルの文字コード: {encoding} (信頼度: {confidence:.2%})")

        # 検出された文字コードでファイルを読み込む
        with open(template_file, 'r', encoding=encoding) as f:
            content = f.read()

        # 件名と本文を分離（最初の行が件名、2行目は空行、3行目以降が本文）
        lines = content.split('\n')

        if len(lines) < 3:
            # i18nがない場合はデフォルトメッセージ（後方互換性のため）
            if i18n:
                raise ValueError("Invalid template format. Line 1: Subject, Line 2: Empty, Line 3+: Body")
            else:
                raise ValueError("テンプレートファイルの形式が正しくありません。1行目: 件名、2行目: 空行、3行目以降: 本文")

        subject = lines[0].strip()
        # 2行目をスキップして3行目以降を本文とする
        body = '\n'.join(lines[2:])

        return subject, body
    
    def create_message(self, to_email, to_name, to_affiliation, subject_template, body_template,
                      cc=None, bcc=None, reply_to=None, attachments=None):
        """
        メールメッセージを作成

        Args:
            to_email: 宛先メールアドレス
            to_name: 宛先氏名
            to_affiliation: 宛先所属名
            subject_template: 件名テンプレート
            body_template: 本文テンプレート
            cc: CCアドレス（カンマ区切りまたはリスト）
            bcc: BCCアドレス（カンマ区切りまたはリスト）
            reply_to: 返信先アドレス
            attachments: 添付ファイルパスのリスト

        Returns:
            MIMEMultipartメッセージオブジェクト
        """
        msg = MIMEMultipart()

        # 送信元の設定（表示名がある場合は formataddr を使用）
        if self.sender_display_name:
            msg['From'] = formataddr((str(Header(self.sender_display_name, 'utf-8')), self.email_address))
        else:
            msg['From'] = self.email_address

        msg['To'] = to_email

        # 件名に所属名と氏名を展開
        subject = subject_template.replace('{所属}', to_affiliation).replace('{氏名}', to_name)
        msg['Subject'] = Header(subject, 'utf-8')

        # CC設定
        if cc:
            if isinstance(cc, str):
                msg['Cc'] = cc
            else:
                msg['Cc'] = ', '.join(cc)

        # BCC設定
        if bcc:
            if isinstance(bcc, str):
                msg['Bcc'] = bcc
            else:
                msg['Bcc'] = ', '.join(bcc)

        # Reply-To設定
        if reply_to:
            msg['Reply-To'] = reply_to

        # 本文に所属名と氏名を展開
        body = body_template.replace('{所属}', to_affiliation).replace('{氏名}', to_name)
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        # 添付ファイルを追加
        if attachments:
            for file_path in attachments:
                if os.path.exists(file_path):
                    with open(file_path, 'rb') as f:
                        # ファイルのMIMEタイプを推測
                        filename = os.path.basename(file_path)
                        mime_type, _ = mimetypes.guess_type(file_path)
                        if mime_type is None:
                            mime_type = 'application/octet-stream'

                        # MIMEタイプを分割（例: 'image/png' -> 'image', 'png'）
                        maintype, subtype = mime_type.split('/', 1)

                        part = MIMEBase(maintype, subtype)
                        part.set_payload(f.read())
                        encoders.encode_base64(part)

                        # Content-Typeヘッダーにnameパラメータを追加（日本語ファイル名対応）
                        encoded_filename = str(Header(filename, 'utf-8'))
                        part.set_param('name', encoded_filename)
                        # Content-Dispositionヘッダーを設定
                        part.add_header('Content-Disposition', 'attachment', filename=encoded_filename)
                        msg.attach(part)

        return msg
    
    def send_bulk_emails(self, csv_file, template_file,
                        cc=None, bcc=None, reply_to=None, attachments=None, delay=1, i18n=None):
        """
        一斉送信を実行

        Args:
            csv_file: 受信者リストCSVファイル
            template_file: メールテンプレートファイル（件名と本文）
            cc: CCアドレス
            bcc: BCCアドレス
            reply_to: 返信先アドレス
            attachments: 添付ファイルパスのリスト
            delay: メール送信間隔（秒）
            i18n: 国際化インスタンス
        """
        # 受信者リストとテンプレートを読み込み
        recipients = self.read_recipients(csv_file, i18n)
        subject_template, body_template = self.read_email_template(template_file, i18n)

        # 確認メッセージ（i18nがある場合は多言語対応、ない場合は日本語デフォルト）
        if i18n:
            print(i18n.get('cli_confirm_header'))
            print(i18n.get('preview_subject', subject_template))
            print(i18n.get('preview_recipients', len(recipients)))
            if self.sender_display_name:
                print(i18n.get('preview_sender', f"{self.sender_display_name} <{self.email_address}>"))
            else:
                print(i18n.get('preview_sender', self.email_address))
            if cc:
                print(i18n.get('preview_cc', cc))
            if bcc:
                print(i18n.get('preview_bcc', bcc))
            if reply_to:
                print(i18n.get('preview_reply_to', reply_to))
            if attachments:
                print(i18n.get('preview_attachments', ', '.join(attachments)))
        else:
            print(f"\n=== 送信内容確認 ===")
            print(f"件名: {subject_template}")
            print(f"送信先: {len(recipients)}件")
            if self.sender_display_name:
                print(f"送信元: {self.sender_display_name} <{self.email_address}>")
            else:
                print(f"送信元: {self.email_address}")
            if cc:
                print(f"CC: {cc}")
            if bcc:
                print(f"BCC: {bcc}")
            if reply_to:
                print(f"Reply-To: {reply_to}")
            if attachments:
                print(f"添付ファイル: {', '.join(attachments)}")

        # 確認
        if i18n:
            confirm = input(i18n.get('cli_confirm_send') + ": ")
        else:
            confirm = input("\n送信を開始しますか？ (yes/no): ")

        if confirm.lower() != 'yes':
            if i18n:
                print(i18n.get('cli_cancelled'))
            else:
                print("送信をキャンセルしました。")
            return
        
        # SMTP接続
        try:
            # ポート465はSSL、それ以外はTLSを使用
            local_hostname = self._get_safe_local_hostname()
            if self.smtp_port == 465:
                server = smtplib.SMTP_SSL(self.smtp_server, self.smtp_port, local_hostname=local_hostname)
            else:
                server = smtplib.SMTP(self.smtp_server, self.smtp_port, local_hostname=local_hostname)
                server.starttls()

            server.login(self.email_address, self.email_password)
            
            # 各受信者にメール送信
            success_count = 0
            fail_count = 0
            
            for i, recipient in enumerate(recipients, 1):
                try:
                    msg = self.create_message(
                        to_email=recipient['email'],
                        to_name=recipient['name'],
                        to_affiliation=recipient['affiliation'],
                        subject_template=subject_template,
                        body_template=body_template,
                        cc=cc,
                        bcc=bcc,
                        reply_to=reply_to,
                        attachments=attachments
                    )

                    # 送信
                    server.send_message(msg)
                    success_count += 1
                    if i18n:
                        print(i18n.get('send_success', i, len(recipients), recipient['affiliation'],
                                      recipient['name'], recipient['email']))
                    else:
                        print(f"[{i}/{len(recipients)}] 送信成功: {recipient['affiliation']} {recipient['name']} ({recipient['email']})")

                    # 送信間隔を設定（レート制限対策）
                    if i < len(recipients):
                        time.sleep(delay)

                except Exception as e:
                    fail_count += 1
                    if i18n:
                        print(i18n.get('send_failed', i, len(recipients), recipient['affiliation'],
                                      recipient['name'], recipient['email'], str(e)))
                    else:
                        print(f"[{i}/{len(recipients)}] 送信失敗: {recipient['affiliation']} {recipient['name']} ({recipient['email']}) - エラー: {e}")

            server.quit()

            if i18n:
                print(f"\n{i18n.get('send_complete', success_count, fail_count)}")
            else:
                print(f"\n送信完了: 成功 {success_count}件, 失敗 {fail_count}件")

        except Exception as e:
            if i18n:
                print(f"SMTP connection error: {e}")
            else:
                print(f"SMTP接続エラー: {e}")


def main():
    """メイン処理"""

    # コマンドライン引数をパース
    parser = argparse.ArgumentParser(description='Email Bulk Sender / メール一括送信ツール')
    parser.add_argument('--lang', choices=['ja', 'en'], help='Language / 言語 (ja/en)')
    parser.add_argument('--load-config', action='store_true', help='Load settings from config file / 設定ファイルから読み込む')
    parser.add_argument('--save-config', action='store_true', help='Save settings to config file / 設定ファイルに保存する')
    args = parser.parse_args()

    # i18nインスタンスを作成
    i18n = get_i18n(args.lang)

    # 設定ファイル管理を初期化
    config_manager = ConfigManager("email")

    # 設定を読み込む（--load-configフラグが指定された場合のみ）
    config = None
    if args.load_config:
        config = config_manager.load_config()
        if config:
            if i18n.get_language() == 'ja':
                print(f"設定ファイルを読み込みました: {config_manager.get_config_path()}\n")
            else:
                print(f"Loaded config from: {config_manager.get_config_path()}\n")
        else:
            if i18n.get_language() == 'ja':
                print(f"警告: 設定ファイルが見つかりません: {config_manager.get_config_path()}\n")
            else:
                print(f"Warning: Config file not found: {config_manager.get_config_path()}\n")

    # 設定がない場合はデフォルトを使用
    if not config:
        config = config_manager.get_default_config()

    # メール設定
    print(i18n.get('cli_title') + "\n")

    # SMTPサーバーの取得（設定ファイルまたはデフォルト値から）
    smtp_server_from_config = config.get('smtp', {}).get('server', '')
    smtp_server = smtp_server_from_config if smtp_server_from_config else DEFAULT_SMTP_SERVER
    if smtp_server:
        if i18n.get_language() == 'ja':
            print(f"SMTPサーバー: {smtp_server} (設定済み)")
        else:
            print(f"SMTP Server: {smtp_server} (configured)")
    else:
        smtp_server = input(i18n.get('cli_smtp_server') + ": ")

    # SMTPポートの取得（設定ファイルまたはデフォルト値から）
    smtp_port_from_config = config.get('smtp', {}).get('port', None)
    smtp_port = None

    # 設定ファイルの値を優先
    if smtp_port_from_config is not None:
        try:
            smtp_port = int(smtp_port_from_config)
        except (ValueError, TypeError):
            if i18n.get_language() == 'ja':
                print(f"警告: 設定されたポート '{smtp_port_from_config}' が無効です。")
            else:
                print(f"Warning: Configured port '{smtp_port_from_config}' is invalid.")
            smtp_port = None

    # DEFAULT値を使用
    if smtp_port is None and DEFAULT_SMTP_PORT:
        try:
            smtp_port = int(DEFAULT_SMTP_PORT)
        except (ValueError, TypeError):
            smtp_port = None

    # 設定済みの値を表示または入力を求める
    if smtp_port is not None:
        if i18n.get_language() == 'ja':
            print(f"SMTPポート: {smtp_port} (設定済み)")
        else:
            print(f"SMTP Port: {smtp_port} (configured)")
    else:
        smtp_port_input = input(i18n.get('cli_smtp_port') + ": ").strip()
        if smtp_port_input:
            try:
                smtp_port = int(smtp_port_input)
            except ValueError:
                if i18n.get_language() == 'ja':
                    print(f"警告: 入力されたポート番号が無効です。デフォルト587を使用します。")
                else:
                    print(f"Warning: Invalid port number. Using default 587.")
                smtp_port = 587
        else:
            smtp_port = 587

    # メールアドレスの取得（設定ファイルまたはデフォルト値から）
    email_from_config = config.get('sender', {}).get('email_address', '')
    email_address = email_from_config if email_from_config else DEFAULT_EMAIL_ADDRESS
    if email_address:
        if i18n.get_language() == 'ja':
            print(f"送信元メールアドレス: {email_address} (設定済み)")
        else:
            print(f"Sender Email: {email_address} (configured)")
    else:
        email_address = input(i18n.get('cli_email_address') + ": ")

    # パスワードの取得（セキュリティのため設定ファイルには保存しない）
    # DEFAULT_EMAIL_PASSWORDがある場合のみ使用（後方互換性のため）
    if DEFAULT_EMAIL_PASSWORD:
        email_password = DEFAULT_EMAIL_PASSWORD
        if i18n.get_language() == 'ja':
            print("メールパスワード: ******** (設定済み)")
        else:
            print("Email Password: ******** (configured)")
    else:
        email_password = getpass(i18n.get('cli_email_password') + ": ")

    # 送信元表示名の取得（設定ファイルまたはデフォルト値から）
    display_name_from_config = config.get('sender', {}).get('display_name', '')
    sender_display_name = display_name_from_config if display_name_from_config else SENDER_DISPLAY_NAME
    if sender_display_name:
        if i18n.get_language() == 'ja':
            print(f"送信元表示名: {sender_display_name} (設定済み)")
        else:
            print(f"Sender Display Name: {sender_display_name} (configured)")
    else:
        sender_display_name = input(i18n.get('cli_sender_name') + ": ").strip()
    
    # ファイルと設定
    csv_from_config = config.get('files', {}).get('csv_file', '')
    csv_file = csv_from_config if csv_from_config else DEFAULT_CSV_FILE
    if csv_file:
        if i18n.get_language() == 'ja':
            print(f"受信者リストCSVファイル: {csv_file} (設定済み)")
        else:
            print(f"CSV File: {csv_file} (configured)")
    else:
        csv_file = input(i18n.get('cli_csv_file') + ": ") or "list.csv"

    template_from_config = config.get('files', {}).get('template_file', '')
    template_file = template_from_config if template_from_config else DEFAULT_TEMPLATE_FILE
    if template_file:
        if i18n.get_language() == 'ja':
            print(f"メールテンプレートファイル: {template_file} (設定済み)")
        else:
            print(f"Template File: {template_file} (configured)")
    else:
        template_file = input(i18n.get('cli_template_file') + ": ") or "body.txt"
    
    # オプション設定
    cc_from_config = config.get('email_options', {}).get('cc', '')
    # 設定ファイルの値が優先、次にDEFAULT値
    if cc_from_config:
        cc = cc_from_config or None
    elif DEFAULT_CC is not None:
        cc = DEFAULT_CC or None
    else:
        cc = None

    if cc or (cc == "" and (cc_from_config or DEFAULT_CC is not None)):
        if i18n.get_language() == 'ja':
            print(f"CC: {cc if cc else 'なし'} (設定済み)")
        else:
            print(f"CC: {cc if cc else 'None'} (configured)")
    else:
        cc = input(i18n.get('cli_cc') + ": ").strip() or None

    bcc_from_config = config.get('email_options', {}).get('bcc', '')
    # 設定ファイルの値が優先、次にDEFAULT値
    if bcc_from_config:
        bcc = bcc_from_config or None
    elif DEFAULT_BCC is not None:
        bcc = DEFAULT_BCC or None
    else:
        bcc = None

    if bcc or (bcc == "" and (bcc_from_config or DEFAULT_BCC is not None)):
        if i18n.get_language() == 'ja':
            print(f"BCC: {bcc if bcc else 'なし'} (設定済み)")
        else:
            print(f"BCC: {bcc if bcc else 'None'} (configured)")
    else:
        bcc = input(i18n.get('cli_bcc') + ": ").strip() or None

    reply_to_from_config = config.get('email_options', {}).get('reply_to', '')
    # 設定ファイルの値が優先、次にDEFAULT値
    if reply_to_from_config:
        reply_to = reply_to_from_config or None
    elif DEFAULT_REPLY_TO is not None:
        reply_to = DEFAULT_REPLY_TO or None
    else:
        reply_to = None

    if reply_to or (reply_to == "" and (reply_to_from_config or DEFAULT_REPLY_TO is not None)):
        if i18n.get_language() == 'ja':
            print(f"Reply-To: {reply_to if reply_to else 'なし'} (設定済み)")
        else:
            print(f"Reply-To: {reply_to if reply_to else 'None'} (configured)")
    else:
        reply_to = input(i18n.get('cli_reply_to') + ": ").strip() or None

    # 添付ファイル設定
    attachments_from_config = config.get('files', {}).get('attachments', [])
    # リストを文字列に変換（設定ファイルではリストとして保存）
    if isinstance(attachments_from_config, list) and attachments_from_config:
        attachments_input = ','.join(attachments_from_config)
    elif isinstance(attachments_from_config, str):
        attachments_input = attachments_from_config
    else:
        attachments_input = DEFAULT_ATTACHMENTS

    if attachments_input:
        if i18n.get_language() == 'ja':
            print(f"添付ファイル: {attachments_input} (設定済み)")
        else:
            print(f"Attachments: {attachments_input} (configured)")
    else:
        if i18n.get_language() == 'ja':
            attachments_input = input("添付ファイル (複数の場合はカンマ区切り、不要ならEnter): ").strip()
        else:
            attachments_input = input("Attachments (comma-separated for multiple, press Enter to skip): ").strip()

    # 添付ファイルのリストを作成し、存在確認
    attachments = None
    if attachments_input:
        attachments = [f.strip() for f in attachments_input.split(',')]
        # ファイルの存在確認
        for file_path in attachments:
            if not os.path.exists(file_path):
                if i18n.get_language() == 'ja':
                    print(f"警告: ファイルが見つかりません: {file_path}")
                    confirm = input("このまま続行しますか？ (yes/no): ")
                else:
                    print(f"Warning: File not found: {file_path}")
                    confirm = input("Continue anyway? (yes/no): ")
                if confirm.lower() != 'yes':
                    print(i18n.get('cli_cancelled'))
                    return

    # 送信遅延時間の取得
    delay_from_config = config.get('email_options', {}).get('send_delay', None)
    send_delay = None

    # 設定ファイルの値を優先
    if delay_from_config is not None:
        try:
            send_delay = float(delay_from_config)
        except (ValueError, TypeError):
            if i18n.get_language() == 'ja':
                print(f"警告: 設定された送信間隔 '{delay_from_config}' が無効です。")
            else:
                print(f"Warning: Configured send_delay '{delay_from_config}' is invalid.")
            send_delay = None

    # DEFAULT値を使用
    if send_delay is None and DEFAULT_SEND_DELAY:
        try:
            send_delay = float(DEFAULT_SEND_DELAY)
        except (ValueError, TypeError):
            send_delay = None

    # 設定済みの値を表示または入力を求める
    if send_delay is not None:
        if i18n.get_language() == 'ja':
            print(f"送信間隔: {send_delay}秒 (設定済み)")
        else:
            print(f"Send Delay: {send_delay} seconds (configured)")
    else:
        if i18n.get_language() == 'ja':
            delay_input = input("送信間隔（秒） (デフォルト: 5): ").strip()
        else:
            delay_input = input("Send Delay (seconds, default: 5): ").strip()
        if delay_input:
            try:
                send_delay = float(delay_input)
            except ValueError:
                if i18n.get_language() == 'ja':
                    print("警告: 入力された値が無効です。デフォルト5秒を使用します。")
                else:
                    print("Warning: Invalid value. Using default 5 seconds.")
                send_delay = 5
        else:
            send_delay = 5

    # 設定を保存（--save-configフラグが指定されている場合）
    if args.save_config:
        # 設定ファイルに保存する内容を構築（パスワードは含めない）
        config_to_save = {
            "version": config_manager.CONFIG_VERSION,
            "smtp": {
                "server": smtp_server,
                "port": smtp_port
            },
            "sender": {
                "email_address": email_address,
                "display_name": sender_display_name
            },
            "files": {
                "csv_file": csv_file,
                "template_file": template_file,
                "attachments": attachments if attachments else []
            },
            "email_options": {
                "cc": cc if cc else "",
                "bcc": bcc if bcc else "",
                "reply_to": reply_to if reply_to else "",
                "send_delay": send_delay
            },
            "ui": {
                "language": i18n.get_language()
            }
        }

        if config_manager.save_config(config_to_save):
            if i18n.get_language() == 'ja':
                print(f"\n設定をファイルに保存しました: {config_manager.get_config_path()}")
            else:
                print(f"\nSaved configuration to: {config_manager.get_config_path()}")
        else:
            if i18n.get_language() == 'ja':
                print("\n警告: 設定ファイルの保存に失敗しました。")
            else:
                print("\nWarning: Failed to save configuration file.")

    # 送信実行
    sender = EmailBulkSender(email_address, email_password, smtp_server, smtp_port, sender_display_name)
    sender.send_bulk_emails(
        csv_file=csv_file,
        template_file=template_file,
        cc=cc,
        bcc=bcc,
        reply_to=reply_to,
        attachments=attachments,
        delay=send_delay,
        i18n=i18n
    )


if __name__ == "__main__":
    main()
